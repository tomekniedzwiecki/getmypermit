"""
Faza 1 Dzien 1: Ekstrakcja unique values ze wszystkich arkuszy Pawla.
Wynik: JSON-y do folderu dane_od_pawla/_preview/_unique/
  - staff.json        (OPIEKUN SPRAWY, Osoba, OPIEKUN, INSPEKTOR - samo z kancelarii)
  - employers.json    (PRACODAWCA, FIRMA)
  - offices.json      (Oddzial, Urzad, miasto z LEGNICA/WALBRZYCH/WROCLAW)
  - inspectors.json   (INSPEKTOR z arkuszy rejestru)
  - stages.json       (ETAP POSTĘPOWANIA)
  - case_types.json   (RODZAJ SPRAWY, Typ sprawy)
  - submission_methods.json (METODA ZŁOŻENIA WNIOSKU)
  - statuses.json     (STATUS SPRAWY)
  - nationalities.json (OBYWATELSTWO)

Pomaga zaplanowac import - kazda wartosc bedzie mapowana do docelowej struktury.
"""
from pathlib import Path
import json
from collections import Counter

# Monkey-patch dla overflow dates
import numbers_parser.cell as _np_cell
from numbers_parser.cell import Cell, TextCell

_original = Cell._from_storage
def _safe(table_id, row, col, buffer, model):
    try:
        return _original(table_id, row, col, buffer, model)
    except OverflowError:
        return TextCell(row, col, "<date_overflow>")
    except Exception as e:
        return TextCell(row, col, f"<err:{type(e).__name__}>")
Cell._from_storage = staticmethod(_safe)

from numbers_parser import Document
from openpyxl import load_workbook

DANE = Path(__file__).parent.parent / "dane_od_pawla"
OUT = DANE / "_preview" / "_unique"
OUT.mkdir(parents=True, exist_ok=True)

import unicodedata
def find_file(fragment: str) -> Path:
    """Znajdz plik po fragmencie nazwy (fuzzy) - radzi sobie z NFC/NFD unicode z Maca."""
    frag_nfc = unicodedata.normalize('NFC', fragment).lower()
    for p in DANE.iterdir():
        name_nfc = unicodedata.normalize('NFC', p.name).lower()
        if frag_nfc in name_nfc:
            return p
    raise FileNotFoundError(f"No file matching fragment '{fragment}' in {DANE}")

def cell_val(cell_or_val):
    try:
        v = cell_or_val.value if hasattr(cell_or_val, 'value') else cell_or_val
        return str(v).strip() if v is not None else ""
    except Exception:
        return ""

def read_numbers_table(path, sheet_name, table_name, header_row=0, data_start_row=1):
    """Czyta tabele z pliku .numbers i zwraca (headers, list of row-dicts)."""
    doc = Document(str(path))
    for sheet in doc.sheets:
        if sheet.name != sheet_name:
            continue
        for table in sheet.tables:
            if table.name != table_name:
                continue
            rows = []
            try:
                n_rows = table.num_rows
                n_cols = table.num_cols
            except Exception:
                return [], []
            headers = []
            for c in range(n_cols):
                try:
                    headers.append(cell_val(table.cell(header_row, c)))
                except Exception:
                    headers.append("")
            for r in range(data_start_row, n_rows):
                row_dict = {}
                for c in range(n_cols):
                    try:
                        row_dict[headers[c] if c < len(headers) else f"col_{c}"] = cell_val(table.cell(r, c))
                    except Exception:
                        pass
                rows.append(row_dict)
            return headers, rows
    return [], []

def read_xlsx_sheet(path, sheet_name):
    wb = load_workbook(str(path), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return [], []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    return rows, rows  # no headers

# ==== ZBIORNIKI ====
collect = {
    "staff": Counter(),             # opiekunowie spraw (kancelaria)
    "employers": Counter(),
    "inspectors": Counter(),
    "offices_raw": Counter(),       # "ŚUW / UDSC", "ŚUW" itp.
    "office_departments": Counter(),
    "stages": Counter(),
    "case_types": Counter(),
    "case_type_notes": Counter(),   # UWAGI DO RODZAJU SPRAWY
    "submission_methods": Counter(),
    "statuses": Counter(),
    "nationalities": Counter(),
    "transport_types": Counter(),
    "forma_platnosci": Counter(),
}

def bump(counter_name, value):
    v = value.strip() if isinstance(value, str) else ""
    if v and v not in ("-", "—", "<date_overflow>"):
        collect[counter_name][v] += 1

# ==== 1. Ewidencja spotkań.numbers ====
print("1/5 Ewidencja spotkań.numbers...")
try:
    headers, rows = read_numbers_table(
        find_file("Ewidencja spotka"),
        "Sprawy kancelaria", "Ewidencja spotkań-1",
        header_row=0, data_start_row=1
    )
    for row in rows:
        bump("staff", row.get("Osoba ", ""))
        bump("case_types", row.get("Typ sprawy", ""))
    print(f"  {len(rows)} wierszy")
except Exception as e:
    print(f"  BLAD: {e}")

# ==== 2. Rejestr- karty pobytu z podziałem.numbers ====
print("2/5 Rejestr- karty pobytu...")
rejestr_tables = [
    ("POBYT", "Tabela 1", 2),           # header_row=2 (pierwsze 2 wiersze to metadane)
    ("POZOSTAŁE", "Tabela 1", 2),
    ("REZYDENT", "Tabela 1", 2),
    ("ZEZWOLENIA TYP A", "Zezwolenia typ A I Zaproszenia", 0),
    ("Odebrane decyzje", "PO WYDANEJ DECYZJI-1", 0),
    ("Smart Work", "SMART WORK", 0),
]
# Actually widzialem ze row0 ma naglowki, row1-2 to metadata. Zmienimy - header_row=0, data_start_row=3
rejestr_tables = [
    ("POBYT", "Tabela 1", 0, 3),
    ("POZOSTAŁE", "Tabela 1", 0, 3),
    ("REZYDENT", "Tabela 1", 0, 3),
    ("ZEZWOLENIA TYP A", "Zezwolenia typ A I Zaproszenia", 0, 1),
    ("Odebrane decyzje", "PO WYDANEJ DECYZJI-1", 0, 1),
    ("Smart Work", "SMART WORK", 0, 1),
]
try:
    for sheet_name, table_name, hdr, data_start in rejestr_tables:
        headers, rows = read_numbers_table(
            find_file("Rejestr- karty pobytu"),
            sheet_name, table_name, hdr, data_start
        )
        print(f"  [{sheet_name}]: {len(rows)} wierszy")
        for row in rows:
            bump("staff", row.get("OPIEKUN SPRAWY", ""))
            bump("employers", row.get("PRACODAWCA", ""))
            bump("inspectors", row.get("INSPEKTOR", ""))
            bump("office_departments", row.get("Oddział", ""))
            bump("stages", row.get("ETAP POSTĘPOWANIA", ""))
            bump("case_types", row.get("RODZAJ SPRAWY", ""))
            bump("case_types", row.get("RODZAJ SPRAWY ", ""))
            bump("case_type_notes", row.get("UWAGI DO RODZAJU SPRAWY", ""))
            bump("submission_methods", row.get("METODA ZŁOŻENIA WNIOSKU", ""))
            bump("statuses", row.get("STATUS SPRAWY", ""))
            # Smart Work
            bump("offices_raw", row.get("Urząd", ""))
except Exception as e:
    import traceback
    traceback.print_exc()

# ==== 3. ROZLICZENIA.xlsx ====
print("3/5 ROZLICZENIA.xlsx...")
try:
    wb = load_workbook(find_file("ROZLICZENIA"), read_only=True, data_only=True)
    ws = wb["ROZLICZENIA - ZESTAWIENIE"]
    for row in ws.iter_rows(values_only=True):
        if row and len(row) > 19:
            # col[3] = typ sprawy, col[5] = forma platnosci, col[19] = miasto/oddzial
            if row[3]: bump("case_types", str(row[3]).strip())
            if row[5]: bump("forma_platnosci", str(row[5]).strip())
            if row[19]: bump("offices_raw", str(row[19]).strip())
except Exception as e:
    print(f"  BLAD: {e}")

# ==== 4. SKŁADANE WNIOSKI.numbers ====
print("4/5 SKŁADANE WNIOSKI.numbers...")
try:
    path = find_file("WNIOSKI")
    doc = Document(str(path))
    for sheet in doc.sheets:
        for table in sheet.tables:
            try:
                n_rows = table.num_rows
                n_cols = table.num_cols
            except Exception:
                continue
            # miasta z nazw arkuszy
            if sheet.name in ("LEGNICA", "WAŁBRZYCH", "WROCŁAW"):
                bump("offices_raw", f"ŚUW {sheet.name}")

            # Czytaj naglowki z row0 (dla wiekszosci) albo z row0/row1 (WALBRZYCH - header w row0 legenda)
            headers = []
            for c in range(n_cols):
                try:
                    headers.append(cell_val(table.cell(0, c)))
                except:
                    headers.append("")

            for r in range(1, n_rows):
                row_dict = {}
                for c in range(n_cols):
                    try:
                        row_dict[headers[c] if c < len(headers) else f"col_{c}"] = cell_val(table.cell(r, c))
                    except Exception:
                        pass
                bump("employers", row_dict.get("PRACODAWCA", ""))
                bump("staff", row_dict.get("OPIEKUN", ""))
                bump("staff", row_dict.get("OPIEKUN SPRAWY", ""))
                bump("nationalities", row_dict.get("OBYWATELSTWO", ""))
                bump("transport_types", row_dict.get("PODRÓŻ", ""))
                bump("transport_types", row_dict.get("Jak jedzie ", ""))
except Exception as e:
    import traceback
    traceback.print_exc()

# ==== ZAPIS ====
print("\n5/5 Zapis wyników...")
for name, counter in collect.items():
    data = {
        "count": len(counter),
        "total_occurrences": sum(counter.values()),
        "values": [{"value": v, "count": c} for v, c in counter.most_common()]
    }
    out_file = OUT / f"{name}.json"
    out_file.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  {name}: {len(counter)} unique -> {out_file.relative_to(DANE.parent)}")

# Summary
print("\n" + "="*70)
print("PODSUMOWANIE UNIQUE VALUES")
print("="*70)
for name, counter in collect.items():
    print(f"\n[{name}] ({len(counter)} unique):")
    for v, c in counter.most_common(10):
        print(f"  {c:4d}x  {v[:60]}")
    if len(counter) > 10:
        print(f"  ... + {len(counter) - 10} wiecej")
