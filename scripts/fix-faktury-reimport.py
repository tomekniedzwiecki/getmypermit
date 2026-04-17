"""
Fix: zaimportuj ponownie FAKTURY DO WYSTAWIENIA z:
- invoice_number (col[4] jesli matchuje A#/#/####)
- status 'issued' jesli col[4] = 'wystawiona'
- issue_date z separatora miesiecznego w col[0]
- zachowaj mapowanie do employers
"""
import os, re, requests, unicodedata
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import load_workbook
load_dotenv(Path(__file__).parent.parent / ".env")

URL = os.environ["SUPABASE_URL"]
KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
H = {"apikey": KEY, "Authorization": f"Bearer {KEY}", "Content-Type": "application/json"}

def find_file(frag):
    d = Path(__file__).parent.parent / "dane_od_pawla"
    for p in d.iterdir():
        if frag.lower() in unicodedata.normalize('NFC', p.name).lower():
            return p

def norm_name(s):
    if not s: return ""
    n = unicodedata.normalize('NFD', s)
    n = ''.join(c for c in n if unicodedata.category(c) != 'Mn')
    return re.sub(r'\s+', ' ', n.lower().strip())

# Pobierz employers do lookup
r = requests.get(f"{URL}/rest/v1/gmp_employers?select=id,name,name_normalized", headers={**H, "Range": "0-9999"})
emps = r.json()
emp_by_name = {e['name_normalized']: e['id'] for e in emps if e.get('name_normalized')}

# Delete old faktury z faktury_do_wystawienia
print("1. Usuwam stare 42 faktury z tego zrodla...")
r = requests.delete(f"{URL}/rest/v1/gmp_invoices?import_source=eq.faktury_do_wystawienia",
                   headers={**H, "Prefer": "return=minimal"})
print(f"   HTTP: {r.status_code}")

# Parse arkusz
wb = load_workbook(str(find_file("ROZLICZENIA")), read_only=True, data_only=True)
ws = wb['FAKTURY DO WYSTAWIENIA']

current_month = None
invoices_to_insert = []

for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
    vals = list(row) + [''] * (6 - len(row))
    date_col = vals[0]
    firma = str(vals[1]).strip() if vals[1] else ''
    kwota = vals[2]
    opis = str(vals[3]).strip() if vals[3] else ''
    col4 = str(vals[4]).strip() if vals[4] else ''
    col5 = str(vals[5]).strip() if vals[5] else ''

    # Separator miesieczny (data w col[0], reszta pusta)
    if date_col and not firma and not kwota and not opis:
        try:
            current_month = date_col if isinstance(date_col, str) else date_col.date().isoformat()
        except:
            current_month = None
        continue

    # Pomin puste
    if not firma and not opis and not kwota:
        continue

    # Parse kwota
    amount = None
    if kwota is not None and kwota != '':
        try:
            amount = float(str(kwota).replace(',', '.').replace(' ', ''))
        except: amount = None
    if not amount:
        # Zostaw z notatka
        amount = 0

    # Employer
    employer_id = None
    if firma:
        employer_id = emp_by_name.get(norm_name(firma))
        if not employer_id:
            # Stworz nowego
            res = requests.post(f"{URL}/rest/v1/gmp_employers",
                              headers={**H, "Prefer": "return=representation"},
                              json={"name": firma, "name_normalized": norm_name(firma)})
            if res.status_code < 300:
                new = res.json()[0] if isinstance(res.json(), list) else res.json()
                employer_id = new['id']
                emp_by_name[norm_name(firma)] = employer_id
            # jesli konflikt unique - znajdz po nazwie
            elif res.status_code == 409:
                lookup = requests.get(f"{URL}/rest/v1/gmp_employers?name_normalized=eq.{norm_name(firma)}&select=id", headers=H).json()
                if lookup: employer_id = lookup[0]['id']

    # Status i numer faktury
    invoice_number = None
    if re.match(r'^[A-Z]+\d+/\d+/\d+', col4):
        invoice_number = col4
        status = 'issued'
    elif col4.lower() == 'wystawiona':
        status = 'issued'
    elif col5:
        status = 'sent'
    else:
        status = 'to_issue'

    # Issue date
    issue_date = current_month
    sent_at = None
    if col5:
        # Extract date "07.01.2026" -> "2026-01-07"
        m = re.search(r'(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})', col5)
        if m:
            d, mo, y = m.groups()
            sent_at = f"{y}-{int(mo):02d}-{int(d):02d}"

    # Sent_to
    sent_to = None
    if col5:
        low = col5.lower()
        if 'julk' in low: sent_to = 'Julka'
        elif 'klient' in low or 'kuk' in low: sent_to = 'Klient/KUK'
        elif 'oleksand' in low: sent_to = 'Oleksandr'
        elif 'kasi' in low: sent_to = 'Kasia'
        elif 'mack' in low: sent_to = 'Macek'
        else: sent_to = col5[:100]

    invoices_to_insert.append({
        "employer_id": employer_id,
        "invoice_number": invoice_number,
        "issue_date": issue_date,
        "amount": amount,
        "description": opis,
        "status": status,
        "sent_to": sent_to,
        "sent_at": sent_at,
        "notes": col5 or None,
        "import_source": "faktury_do_wystawienia",
        "import_raw": {"raw_row_idx": r_idx, "col4_raw": col4, "col5_raw": col5},
    })

print(f"\n2. Sparsowano {len(invoices_to_insert)} faktur")

# Statystyki
from collections import Counter
stats_status = Counter(i['status'] for i in invoices_to_insert)
print(f"   Po statusie:")
for s, n in stats_status.most_common():
    print(f"     {s}: {n}")
print(f"   Z invoice_number: {sum(1 for i in invoices_to_insert if i.get('invoice_number'))}")
print(f"   Z issue_date: {sum(1 for i in invoices_to_insert if i.get('issue_date'))}")
print(f"   Z employer_id: {sum(1 for i in invoices_to_insert if i.get('employer_id'))}")

# Normalizuj klucze
all_keys = set()
for r in invoices_to_insert: all_keys.update(r.keys())
invoices_to_insert = [{k: r.get(k) for k in all_keys} for r in invoices_to_insert]

# Insert batch
print("\n3. Wstawiam...")
BATCH = 100
inserted = 0
for i in range(0, len(invoices_to_insert), BATCH):
    batch = invoices_to_insert[i:i+BATCH]
    r = requests.post(f"{URL}/rest/v1/gmp_invoices",
                     headers={**H, "Prefer": "return=minimal"},
                     json=batch)
    if r.status_code < 300:
        inserted += len(batch)
    else:
        print(f"   BLAD: {r.status_code} {r.text[:200]}")

print(f"\n=== DONE: {inserted} faktur ===")

# Weryfikacja
r = requests.get(f"{URL}/rest/v1/gmp_invoices?select=count&import_source=eq.faktury_do_wystawienia",
                headers={**H, "Prefer": "count=exact", "Range": "0-0"})
print(f"W bazie: {r.headers.get('content-range', '').split('/')[-1]}")
