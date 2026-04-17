"""
Import pipeline: dane Pawla -> nowy Supabase CRM.

Tryby:
  py scripts/phase1-import.py --dry-run  (default)
     Buduje model w pamieci, zapisuje statystyki do _preview/_dry_run/
     NIC nie trafia do Supabase.

  py scripts/phase1-import.py --live
     Wysyla do Supabase przez service_role REST API.
     Wymaga potwierdzenia przed uruchomieniem.

Kolejnosc:
  1. Slowniki: offices, departments, staff, inspectors
  2. Employers (unique z wszystkich arkuszy)
  3. Clients (z dedupem)
  4. Cases: rejestr (POBYT/POZOSTALE/REZYDENT) + zezwolenia + smart_work + ewidencja + rozliczenia
  5. Case activities: odebrane decyzje (po match z cases)
  6. Payments + payment_plans + invoices
  7. Submissions_queue + appointments (osobiste_odciski)
"""
import argparse
import json
import os
import re
import sys
import unicodedata
from collections import defaultdict, Counter
from datetime import datetime, date
from pathlib import Path
from dotenv import load_dotenv

# Monkey-patch dla overflow dates
import numbers_parser.cell as _np_cell
from numbers_parser.cell import Cell, TextCell
_orig_from_storage = Cell._from_storage
def _safe_from_storage(table_id, row, col, buffer, model):
    try:
        return _orig_from_storage(table_id, row, col, buffer, model)
    except OverflowError:
        return TextCell(row, col, "<date_overflow>")
    except Exception as e:
        return TextCell(row, col, f"<err:{type(e).__name__}>")
Cell._from_storage = staticmethod(_safe_from_storage)

from numbers_parser import Document
from openpyxl import load_workbook

# ==============================================================
# CONFIG
# ==============================================================
ROOT = Path(__file__).parent.parent
DANE = ROOT / "dane_od_pawla"
MAPPINGS = DANE / "_preview" / "_mappings"
DRY_RUN_OUT = DANE / "_preview" / "_dry_run"
DRY_RUN_OUT.mkdir(parents=True, exist_ok=True)

load_dotenv(ROOT / ".env")


# ==============================================================
# NORMALIZACJA / HELPERS
# ==============================================================
def norm_ascii(s: str) -> str:
    """Lowercase + strip + remove Polish diacritics. Dla dedupu."""
    if not s:
        return ""
    n = unicodedata.normalize('NFD', s)
    n = ''.join(c for c in n if unicodedata.category(c) != 'Mn')
    return n.lower().strip()


def norm_name(s: str) -> str:
    """Do dedupu klientow/pracodawcow - lower + no diacritics + collapse whitespace."""
    n = norm_ascii(s)
    return re.sub(r'\s+', ' ', n)


def parse_date(value):
    """Parsuje date z roznych formatow. Zwraca datetime.date lub None."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if s in ("", "-", "<date_overflow>", "BRAK") or s.startswith("<"):
        return None
    # Typowe formaty
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            d = datetime.strptime(s, fmt).date()
            # Odrzucamy niesensowne daty (<1990 lub >2100)
            if 1990 <= d.year <= 2100:
                return d
            return None
        except ValueError:
            continue
    return None


def parse_amount(value):
    """Parsuje kwote. '2000', '200 EURO', 'BRAK' -> float/None."""
    if value is None or value == "":
        return None
    s = str(value).strip().replace(",", ".").replace(" ", "").replace("\xa0", "")
    if s.lower() in ("", "brak", "-", "x"):
        return None
    # Wyciagnij pierwsza liczbe
    m = re.search(r'\d+(\.\d+)?', s)
    if m:
        try:
            return float(m.group())
        except ValueError:
            return None
    return None


def parse_phone(value):
    """Normalizacja numeru telefonu: tylko cyfry."""
    if not value:
        return None
    s = str(value).strip()
    # Usun kropki, nietypowe oznaczenia
    digits = re.sub(r'\D', '', s)
    if len(digits) < 6:
        return None
    # Pierwsze 9 cyfr to numer jesli za duzo (polskie formaty)
    if digits.startswith('48') and len(digits) == 11:
        return digits[2:]
    if digits.endswith('.0'):  # openpyxl czasem zwraca jako float
        digits = digits[:-2]
    return digits[-9:] if len(digits) >= 9 else digits


def cell_val(cell):
    try:
        v = cell.value if hasattr(cell, 'value') else cell
        if v is None:
            return ""
        return str(v).strip()
    except Exception:
        return ""


def find_file(fragment: str) -> Path:
    frag = unicodedata.normalize('NFC', fragment).lower()
    for p in DANE.iterdir():
        name_nfc = unicodedata.normalize('NFC', p.name).lower()
        if frag in name_nfc:
            return p
    raise FileNotFoundError(fragment)


# ==============================================================
# LOAD MAPPINGS
# ==============================================================
print("Ladowanie mappingow...")
staff_mapping = json.loads((MAPPINGS / "staff.json").read_text(encoding="utf-8"))
stages_mapping = {m["source_value"]: m["mapped_stage"] for m in json.loads((MAPPINGS / "stages.json").read_text(encoding="utf-8"))["mapping"]}
status_mapping = json.loads((MAPPINGS / "statuses.json").read_text(encoding="utf-8"))
submission_methods_mapping = json.loads((MAPPINGS / "submission_methods.json").read_text(encoding="utf-8"))
offices_data = json.loads((MAPPINGS / "offices.json").read_text(encoding="utf-8"))
inspectors_mapping = json.loads((MAPPINGS / "inspectors.json").read_text(encoding="utf-8"))
staff_alias_lookup = {k: v for k, v in staff_mapping["alias_lookup"].items()}

BAD_INSPECTOR_RE = re.compile(
    r"^(pc\s*\d+|oc\s?ii|oc\s?\d+|ocii|due|op|\?+|\s*)$",
    re.IGNORECASE
)


def lookup_staff(name_raw: str) -> str | None:
    """Zwraca canonical full_name lub None (do pominiecia)."""
    if not name_raw:
        return None
    key = norm_ascii(name_raw).upper()
    # Sprawdz bezposredni
    for alias_norm, full in staff_alias_lookup.items():
        if alias_norm == key:
            return full
    # Sprawdz czy alias jest prefixem/zawarty
    for alias_norm, full in staff_alias_lookup.items():
        if key and alias_norm and (key in alias_norm or alias_norm in key):
            return full
    return None


# ==============================================================
# MODEL W PAMIECI (in-memory staging)
# ==============================================================
class ImportModel:
    def __init__(self):
        # Slowniki
        self.offices = []                   # list of {name, city, code}
        self.office_departments = []        # list of {office_code, code, name}
        self.inspectors = {}                # norm_name -> {full_name, full_name_normalized}
        self.staff = []                     # list of {full_name, role, aliases}

        # Dane
        self.employers = {}                 # norm_name -> {name, name_normalized, occurrences}
        self.clients = {}                   # (norm_name, birth_date) or (norm_name, None, unique_id) -> client dict
        self.cases = []                     # list of case dicts with refs (client_key, employer_key, staff_name, inspector_name, office_code, dept_code)
        self.activities = []                # activity entries (including decyzje)
        self.orphan_decisions = []          # decyzje bez match
        self.payments = []
        self.payment_plans = []
        self.invoices = []
        self.submissions_queue = []
        self.appointments = []

        # Statystyki
        self.stats = defaultdict(lambda: defaultdict(int))
        self.warnings = []

    def warn(self, category, msg):
        self.warnings.append({"category": category, "msg": msg})
        self.stats["warnings"][category] += 1

    def upsert_employer(self, name: str) -> str | None:
        """Zwraca klucz (norm name) employer lub None jesli pusty."""
        if not name:
            return None
        nn = norm_name(name)
        if not nn:
            return None
        if nn not in self.employers:
            self.employers[nn] = {
                "name": name.strip(),
                "name_normalized": nn,
                "occurrences": 0,
            }
        self.employers[nn]["occurrences"] += 1
        return nn

    def upsert_client(self, last_name: str, first_name: str, birth_date=None, phone=None, email=None, nationality=None, employer_key=None):
        """Dedup. Zwraca unikalny klucz klienta."""
        if not last_name and not first_name:
            return None
        ln = last_name.strip() if last_name else ""
        fn = first_name.strip() if first_name else ""
        full_norm = norm_name(f"{ln} {fn}")
        if not full_norm:
            return None

        if birth_date:
            key = ("WITH_DOB", full_norm, birth_date.isoformat())
        else:
            # Bez daty - osobne rekordy (decyzja D2)
            key = ("NO_DOB", full_norm, len(self.clients))

        if key not in self.clients:
            self.clients[key] = {
                "key": key,
                "last_name": ln,
                "first_name": fn,
                "full_name_normalized": full_norm,
                "birth_date": birth_date.isoformat() if birth_date else None,
                "phone": phone,
                "email": email,
                "nationality": nationality,
                "employer_key": employer_key,
                "occurrences": 0,
            }
        c = self.clients[key]
        c["occurrences"] += 1
        # Uzupelnij jesli wczesniej brakowalo
        if phone and not c.get("phone"):
            c["phone"] = phone
        if email and not c.get("email"):
            c["email"] = email
        if nationality and not c.get("nationality"):
            c["nationality"] = nationality
        if employer_key and not c.get("employer_key"):
            c["employer_key"] = employer_key
        return key

    def upsert_inspector(self, name: str) -> str | None:
        if not name:
            return None
        s = name.strip()
        if BAD_INSPECTOR_RE.match(s):
            return None
        nn = norm_name(s)
        if nn not in self.inspectors:
            self.inspectors[nn] = {"full_name": s, "full_name_normalized": nn}
        return nn

    def find_client_by_name(self, last_name: str, first_name: str, birth_date=None):
        """Szuka istniejacego klienta po nazwisko+imie (+opt. data urodzenia)."""
        full_norm = norm_name(f"{last_name} {first_name}")
        if birth_date:
            key = ("WITH_DOB", full_norm, birth_date.isoformat())
            if key in self.clients:
                return self.clients[key]
        # Bez daty - sprawdz wszystkie wpisy NO_DOB + WITH_DOB z tym imieniem
        matches = [c for k, c in self.clients.items() if c["full_name_normalized"] == full_norm]
        if len(matches) == 1:
            return matches[0]
        # Wiele matchy - ambiguous
        return None


model = ImportModel()


# ==============================================================
# BUDOWA SLOWNIKOW
# ==============================================================
print("\n=== SLOWNIKI ===")
# Offices
for office in offices_data["offices"]:
    model.offices.append(office)
for dept in offices_data["departments"]:
    model.office_departments.append(dept)
print(f"  offices: {len(model.offices)}")
print(f"  office_departments: {len(model.office_departments)}")

# Staff
for person in staff_mapping["staff"]:
    model.staff.append({
        "full_name": person["full_name"],
        "role": person["role"],
        "aliases": person["aliases"],
        "email": person.get("email"),
    })
print(f"  staff: {len(model.staff)}")


# ==============================================================
# IMPORT: REJESTR POBYT / POZOSTAŁE / REZYDENT
# ==============================================================
def import_rejestr_sheet(sheet_name: str, table_name: str, category: str, has_submission_method: bool):
    path = find_file("Rejestr- karty pobytu")
    doc = Document(str(path))
    for sheet in doc.sheets:
        if sheet.name != sheet_name:
            continue
        for table in sheet.tables:
            if table.name != table_name:
                continue
            try:
                n_rows = table.num_rows
                n_cols = table.num_cols
            except Exception as e:
                model.warn("table_read", f"{sheet_name}/{table_name}: {e}")
                return

            # Headers w row 0 (pierwsze 2 wiersze to legenda, dane od row 3)
            headers = [cell_val(table.cell(0, c)) for c in range(n_cols)]

            for r in range(3, n_rows):
                row = {}
                for c in range(n_cols):
                    h = headers[c] if c < len(headers) and headers[c] else f"col_{c}"
                    row[h] = cell_val(table.cell(r, c))

                # Pomin puste wiersze (brak nazwiska)
                last_name = row.get("NAZWISKO", "")
                first_name = row.get("IMIĘ", "")
                if not last_name and not first_name:
                    model.stats[sheet_name]["skipped_empty"] += 1
                    continue

                # Build client
                birth = parse_date(row.get("DATA URODZENIA", ""))
                phone = parse_phone(row.get("KONTAKT TELEFON", ""))
                email = row.get("KONTAKT EMAIL", "") or row.get("KONTAKT EMAIL ", "")
                employer_name = row.get("PRACODAWCA", "")

                employer_key = model.upsert_employer(employer_name)
                client_key = model.upsert_client(
                    last_name, first_name,
                    birth_date=birth,
                    phone=phone,
                    email=email if email else None,
                    employer_key=employer_key,
                )

                if not client_key:
                    model.stats[sheet_name]["skipped_no_client"] += 1
                    continue

                # Status
                raw_status = row.get("STATUS SPRAWY", "").strip()
                status = status_mapping.get(raw_status, "aktywna")

                # Stage
                raw_stage = row.get("ETAP POSTĘPOWANIA", "").strip()
                stage = stages_mapping.get(raw_stage) if raw_stage else None
                stage_notes = raw_stage if raw_stage and not stage else None

                # Metoda zlozenia
                submission_method = None
                if has_submission_method:
                    raw_sm = row.get("METODA ZŁOŻENIA WNIOSKU", "").strip()
                    submission_method = submission_methods_mapping.get(raw_sm)

                # Rodzaj sprawy
                case_type = row.get("RODZAJ SPRAWY", "") or row.get("RODZAJ SPRAWY ", "")
                case_type_notes = row.get("UWAGI DO RODZAJU SPRAWY", "")

                # Kind
                kind = "nowa_sprawa"
                if case_type_notes and "PRZYŁĄCZENIE" in case_type_notes.upper():
                    kind = "przystapienie_do_sprawy"

                # Daty
                date_accepted = parse_date(row.get("DATA PRZYJĘCIA SPRAWY", ""))
                date_submitted = parse_date(row.get("DATA ZŁOŻENIA WNIOSKU", ""))
                date_joined = parse_date(row.get("DATA PRZYSTĄPIENIA DO SPRAWY (SPRAWY Z PRZYSTĄPIENIA)", ""))
                date_transfer_request = parse_date(row.get("DATA WNIOSKU O PRZEKAZANIE DO INNEGO URZĘDU (JEŻELI DOTYCZY)", ""))
                date_transferred = parse_date(row.get("DATA PRZEKAZANIA (JEŻELI DOTYCZY)", ""))

                # Inspektor
                inspector_key = model.upsert_inspector(row.get("INSPEKTOR", ""))

                # Oddzial
                dept_code = row.get("Oddział", "").strip() or None

                # Staff
                opiekun = row.get("OPIEKUN SPRAWY", "").strip()
                staff_name = lookup_staff(opiekun)
                if opiekun and not staff_name:
                    model.stats[sheet_name]["unknown_opiekun"] += 1

                # Finanse
                fee = parse_amount(row.get("WYNAGRODZENIE", ""))
                fee_notes = row.get("UWAGI DO WYNAGRODZENIA", "")
                paragon_raw = row.get("PARAGON", "").lower()
                paragon = paragon_raw in ("tak", "yes", "true", "x")

                case = {
                    "case_number": row.get("NUMER SPRAWY ", "").strip() or row.get("NUMER SPRAWY", "").strip() or None,
                    "znak_sprawy": row.get("ZNAK SPRAWY", "").strip() or None,
                    "client_key": client_key,
                    "employer_key": employer_key,
                    "staff_name": staff_name,
                    "inspector_key": inspector_key,
                    "office_code": "DUW_WROCLAW",  # domyslnie
                    "department_code": dept_code,
                    "status": status,
                    "status_notes": row.get(" UWAGI DO STATUSU", "") or row.get("UWAGI DO STATUSU", ""),
                    "stage": stage,
                    "stage_notes": stage_notes,
                    "kind": kind,
                    "case_type": case_type.strip() if case_type else None,
                    "case_type_notes": case_type_notes.strip() if case_type_notes else None,
                    "submission_method": submission_method,
                    "category": category,
                    "date_accepted": date_accepted.isoformat() if date_accepted else None,
                    "date_submitted": date_submitted.isoformat() if date_submitted else None,
                    "date_joined": date_joined.isoformat() if date_joined else None,
                    "date_transfer_request": date_transfer_request.isoformat() if date_transfer_request else None,
                    "date_transferred": date_transferred.isoformat() if date_transferred else None,
                    "fee_amount": fee,
                    "fee_notes": fee_notes or None,
                    "paragon": paragon,
                    "extra_notes": row.get("DODATKOWY KOMENTARZ ", "") or row.get("DODATKOWY KOMENTARZ", "") or None,
                    "import_source": f"rejestr_{category}",
                    "import_source_row": r,
                    "import_raw": row,
                }
                model.cases.append(case)
                model.stats[sheet_name]["imported"] += 1


print("\n=== REJESTR ===")
for sheet, cat, has_sm in [
    ("POBYT", "pobyt", False),          # POBYT nie ma METODA ZŁOŻENIA
    ("POZOSTAŁE", "pozostale", True),
    ("REZYDENT", "rezydent", True),
]:
    print(f"  [{sheet}] importowanie...")
    import_rejestr_sheet(sheet, "Tabela 1", cat, has_sm)
    print(f"    imported: {model.stats[sheet]['imported']}, skipped_empty: {model.stats[sheet]['skipped_empty']}")


# ==============================================================
# IMPORT: ZEZWOLENIA TYP A
# ==============================================================
def import_zezwolenia_a():
    path = find_file("Rejestr- karty pobytu")
    doc = Document(str(path))
    for sheet in doc.sheets:
        if sheet.name != "ZEZWOLENIA TYP A":
            continue
        for table in sheet.tables:
            if table.name != "Zezwolenia typ A I Zaproszenia":
                continue
            try:
                n_rows = table.num_rows
                n_cols = table.num_cols
            except Exception:
                return
            headers = [cell_val(table.cell(0, c)) for c in range(n_cols)]
            for r in range(1, n_rows):
                row = {headers[c] if headers[c] else f"col_{c}": cell_val(table.cell(r, c)) for c in range(n_cols)}
                firma = row.get("FIRMA", "").strip()
                cudz = row.get("CUDZOZIEMIEC", "").strip()
                if not firma and not cudz:
                    model.stats["ZEZWOLENIA TYP A"]["skipped_empty"] += 1
                    continue
                # cudzoziemiec moze byc "IMIĘ NAZWISKO" albo "NAZWISKO IMIĘ" - zakladamy ze najpierw nazwisko
                parts = cudz.split()
                last = parts[0] if parts else ""
                first = " ".join(parts[1:]) if len(parts) > 1 else ""

                employer_key = model.upsert_employer(firma) if firma else None
                client_key = model.upsert_client(last, first, employer_key=employer_key) if (last or first) else None

                insp = row.get("INSPEKTOR", "") or row.get("col_7", "")
                inspector_key = model.upsert_inspector(insp)

                date_sub = parse_date(row.get("DATA ZŁOŻENIA", ""))
                date_pickup = parse_date(row.get("DATA ODBIORU ", "") or row.get("DATA ODBIORU", ""))
                date_sent = parse_date(row.get("DATA WYDANIA/WYSŁANIA KLIENTOWI", ""))

                case = {
                    "client_key": client_key,
                    "employer_key": employer_key,
                    "inspector_key": inspector_key,
                    "office_code": "DUW_WROCLAW",
                    "status": "zakonczona" if date_sent else "aktywna",
                    "stage": "zakonczenie" if date_sent else None,
                    "case_type": "ZEZWOLENIE TYP A",
                    "category": "zezwolenie_a",
                    "date_submitted": date_sub.isoformat() if date_sub else None,
                    "extra_notes": row.get("BRAKI", "") or None,
                    "import_source": "rejestr_zezwolenia_a",
                    "import_source_row": r,
                    "import_raw": row,
                }
                # Dorzuc dodatkowe info do notes
                extras = []
                if date_pickup:
                    extras.append(f"DATA ODBIORU: {date_pickup.isoformat()}")
                if date_sent:
                    extras.append(f"WYDANIE/WYSŁANIE: {date_sent.isoformat()}")
                if extras:
                    case["extra_notes"] = (case.get("extra_notes") or "") + " | " + " | ".join(extras)

                model.cases.append(case)
                model.stats["ZEZWOLENIA TYP A"]["imported"] += 1

print(f"\n  [ZEZWOLENIA TYP A] importowanie...")
import_zezwolenia_a()
print(f"    imported: {model.stats['ZEZWOLENIA TYP A']['imported']}")


# ==============================================================
# IMPORT: SMART WORK
# ==============================================================
def import_smart_work():
    path = find_file("Rejestr- karty pobytu")
    doc = Document(str(path))
    for sheet in doc.sheets:
        if sheet.name != "Smart Work":
            continue
        for table in sheet.tables:
            if table.name != "SMART WORK":
                continue
            try:
                n_rows = table.num_rows
                n_cols = table.num_cols
            except Exception:
                return
            headers = [cell_val(table.cell(0, c)) for c in range(n_cols)]
            for r in range(1, n_rows):
                row = {headers[c] if headers[c] else f"col_{c}": cell_val(table.cell(r, c)) for c in range(n_cols)}
                last = row.get("Nazwisko", "").strip()
                first = row.get("Imię", "").strip()
                if not last and not first:
                    continue
                firma = row.get("Spółka", "").strip()
                employer_key = model.upsert_employer(firma) if firma else None
                client_key = model.upsert_client(last, first, employer_key=employer_key)
                etap = row.get("Etap", "").strip()
                status = "zakonczona" if etap.upper() == "ZAKOŃCZONE" else "aktywna"

                case = {
                    "client_key": client_key,
                    "employer_key": employer_key,
                    "znak_sprawy": row.get("Sygnatura", "").strip() or None,
                    "status": status,
                    "category": "smart_work",
                    "case_type": "SMART WORK",
                    "extra_notes": f"Urząd: {row.get('Urząd', '')}" if row.get("Urząd") else None,
                    "import_source": "rejestr_smart_work",
                    "import_source_row": r,
                    "import_raw": row,
                }
                model.cases.append(case)
                model.stats["Smart Work"]["imported"] += 1

print(f"\n  [Smart Work] importowanie...")
import_smart_work()
print(f"    imported: {model.stats['Smart Work']['imported']}")


# ==============================================================
# IMPORT: ODEBRANE DECYZJE -> activities (match) lub orphan_decisions
# ==============================================================
def import_odebrane_decyzje():
    path = find_file("Rejestr- karty pobytu")
    doc = Document(str(path))
    for sheet in doc.sheets:
        if sheet.name != "Odebrane decyzje":
            continue
        for table in sheet.tables:
            if table.name != "PO WYDANEJ DECYZJI-1":
                continue
            try:
                n_rows = table.num_rows
                n_cols = table.num_cols
            except Exception:
                return
            headers = [cell_val(table.cell(0, c)) for c in range(n_cols)]
            for r in range(1, n_rows):
                row = {headers[c] if headers[c] else f"col_{c}": cell_val(table.cell(r, c)) for c in range(n_cols)}
                full_name = row.get("NAZWISKO I IMIĘ", "").strip()
                birth = parse_date(row.get("DATA URODZENIA", ""))
                if not full_name:
                    continue
                # Split "NAZWISKO IMIĘ" - zakladamy pierwszy slowo to nazwisko
                parts = full_name.split()
                last = parts[0] if parts else ""
                first = " ".join(parts[1:]) if len(parts) > 1 else ""

                date_received = parse_date(row.get("DATA WPŁYWU DECYZJI", ""))
                date_issued = parse_date(row.get("DATA WYDANEJ DECYZJI", ""))
                delivered = row.get("WYDANIE DECYZJI KLIENTOWI", "")
                date_delivered = parse_date(delivered) if delivered else None

                # Probujemy znalezc istniejacego klienta
                matched_client = model.find_client_by_name(last, first, birth)

                activity = {
                    "activity_type": "decision_received",
                    "content": f"Decyzja odebrana (import historyczny). Klient: {full_name}",
                    "metadata": {
                        "date_issued": date_issued.isoformat() if date_issued else None,
                        "date_received": date_received.isoformat() if date_received else None,
                        "date_delivered": date_delivered.isoformat() if date_delivered else None,
                        "delivery_note": delivered if not date_delivered else None,
                    },
                    "import_source": "rejestr_odebrane_decyzje",
                    "import_source_row": r,
                }

                if matched_client:
                    # Dolaczamy do pierwszej (najstarszej) sprawy klienta
                    activity["_match_client_key"] = matched_client["key"]
                    model.activities.append(activity)
                    model.stats["Odebrane decyzje"]["matched"] += 1
                else:
                    orphan = {
                        "raw_full_name": full_name,
                        "raw_birth_date": birth.isoformat() if birth else None,
                        "date_issued": date_issued.isoformat() if date_issued else None,
                        "date_received": date_received.isoformat() if date_received else None,
                        "date_delivered_to_client": date_delivered.isoformat() if date_delivered else None,
                        "extra_notes": delivered if not date_delivered else None,
                    }
                    model.orphan_decisions.append(orphan)
                    model.stats["Odebrane decyzje"]["orphan"] += 1

print(f"\n  [Odebrane decyzje] importowanie...")
import_odebrane_decyzje()
print(f"    matched: {model.stats['Odebrane decyzje']['matched']}, orphan: {model.stats['Odebrane decyzje']['orphan']}")


# ==============================================================
# IMPORT: EWIDENCJA SPOTKAŃ (leady)
# ==============================================================
def import_ewidencja():
    path = find_file("Ewidencja spotka")
    doc = Document(str(path))
    for sheet in doc.sheets:
        for table in sheet.tables:
            if table.name != "Ewidencja spotkań-1":
                continue
            try:
                n_rows = table.num_rows
                n_cols = table.num_cols
            except Exception:
                return
            headers = [cell_val(table.cell(0, c)) for c in range(n_cols)]
            for r in range(1, n_rows):
                row = {headers[c] if headers[c] else f"col_{c}": cell_val(table.cell(r, c)) for c in range(n_cols)}
                first = row.get("Imię", "").strip()
                last = row.get("Nazwisko", "").strip()
                if not first and not last:
                    continue
                client_key = model.upsert_client(last, first)

                date_meeting = parse_date(row.get("Data", ""))
                przyjeta = row.get("Przyjęta", "").strip().upper()
                status = "zlecona" if przyjeta == "TAK" else "lead"

                opiekun = row.get("Osoba ", "").strip()
                staff_name = lookup_staff(opiekun)

                faktura_raw = row.get("Faktura", "").strip().lower()

                case = {
                    "client_key": client_key,
                    "staff_name": staff_name,
                    "status": status,
                    "case_type": row.get("Typ sprawy", "").strip() or None,
                    "category": "lead_ewidencja",
                    "date_accepted": date_meeting.isoformat() if date_meeting and status == "zlecona" else None,
                    "fee_notes": row.get("Opłata", "") or None,
                    "extra_notes": row.get("Informacja dodatkowa", "") or None,
                    "paragon": "paragon" in (row.get("Opłata", "") or "").lower(),
                    "import_source": "ewidencja_spotkan",
                    "import_source_row": r,
                    "import_raw": row,
                }
                model.cases.append(case)
                model.stats["Ewidencja spotkań"]["imported"] += 1

                # Osobny appointment (konsultacja)
                if date_meeting:
                    model.appointments.append({
                        "appointment_type": "konsultacja",
                        "client_key": client_key,
                        "staff_name": staff_name,
                        "_link_case_idx": len(model.cases) - 1,
                        "scheduled_date": date_meeting.isoformat(),
                        "import_source": "ewidencja_spotkan",
                    })

                # Faktura
                if faktura_raw in ("tak", "true"):
                    model.invoices.append({
                        "_link_case_idx": len(model.cases) - 1,
                        "client_key": client_key,
                        "amount": 0,  # nieznana
                        "description": f"Faktura z ewidencji: {first} {last}",
                        "status": "to_issue",
                        "import_source": "ewidencja_spotkan",
                    })

print(f"\n  [Ewidencja spotkań] importowanie...")
import_ewidencja()
print(f"    imported: {model.stats['Ewidencja spotkań']['imported']}")


# ==============================================================
# IMPORT: ROZLICZENIA.xlsx
# ==============================================================
# KOLUMNY (brak naglowkow - dedukowane):
# 0=data_przyjecia, 1=nazwisko, 2=imie, 3=typ_sprawy, 4=podstawa,
# 5=forma_platnosci, 6=kwota_calk, 7=rata1_kwota, 8=rata1_data,
# 9=rata2_kwota, 10=rata2_data, 11-17=checkboxy X (D4: save raw),
# 18=uwagi, 19=oddzial, 20-22=puste
def import_rozliczenia():
    wb = load_workbook(find_file("ROZLICZENIA"), read_only=True, data_only=True)
    ws = wb["ROZLICZENIA - ZESTAWIENIE"]
    for r, row in enumerate(ws.iter_rows(values_only=True)):
        if not row or all(v is None or v == "" for v in row):
            continue
        row = list(row) + [""] * max(0, 23 - len(row))  # padding

        date_accepted = parse_date(row[0])
        last_name = str(row[1]).strip() if row[1] else ""
        first_name = str(row[2]).strip() if row[2] else ""

        if not last_name and not first_name:
            model.stats["ROZLICZENIA"]["skipped_empty"] += 1
            continue

        client_key = model.upsert_client(last_name, first_name)

        case_type = str(row[3]).strip() if row[3] else None
        case_type_notes = str(row[4]).strip() if row[4] else None  # "KARTA POBYTU"

        forma_raw = str(row[5]).strip().upper() if row[5] else ""
        # Jesli forma zaczyna sie od FAKTURA - to faktura (na employera)
        is_invoice = forma_raw.startswith("FAKTURA")
        employer_key = None
        if is_invoice:
            # Wyciagnij firme po "FAKTURA "
            employer_candidate = forma_raw.replace("FAKTURA", "", 1).strip()
            if employer_candidate:
                employer_key = model.upsert_employer(employer_candidate)
            payment_method = "faktura"
        elif "GOTÓWKA" in forma_raw or "GOTOWKA" in forma_raw:
            payment_method = "gotowka"
        elif "PRZELEW" in forma_raw:
            payment_method = "przelew"
        elif "KARTA" in forma_raw:
            payment_method = "karta"
        else:
            payment_method = "inne" if forma_raw else None

        total_amount = parse_amount(row[6])
        rata1_amount = parse_amount(row[7])
        rata1_date = parse_date(row[8])
        rata2_amount = parse_amount(row[9])
        rata2_date = parse_date(row[10])

        # Checkboxy 11-17 (D4)
        checkboxes = {f"col_{i}": str(row[i]).strip() if row[i] else "" for i in range(11, 18)}

        uwagi = str(row[18]).strip() if row[18] else None
        oddzial = str(row[19]).strip() if row[19] else None

        # Utworz case
        case_idx = len(model.cases)
        case = {
            "client_key": client_key,
            "employer_key": employer_key,
            "status": "aktywna",  # z rozliczen nie wiemy - przyjmujemy aktywna
            "case_type": case_type,
            "case_type_notes": case_type_notes,
            "category": "rozliczenie_historyczne",
            "date_accepted": date_accepted.isoformat() if date_accepted else None,
            "fee_amount": total_amount,
            "fee_notes": f"Forma: {forma_raw}" if forma_raw else None,
            "extra_notes": uwagi,
            "import_source": "rozliczenia",
            "import_source_row": r,
            "import_raw": {
                "raw_row": [str(v) if v is not None else "" for v in row[:23]],
                "checkboxes_11_to_17": checkboxes,
                "oddzial": oddzial,
            },
        }
        model.cases.append(case)

        # Payment plan + payments
        if total_amount:
            model.payment_plans.append({
                "_link_case_idx": case_idx,
                "total_amount": total_amount,
                "payer_type": "employer" if employer_key else "client",
                "client_key": client_key if not employer_key else None,
                "employer_key": employer_key,
                "installments_planned": 2 if rata2_amount else 1,
                "import_source": "rozliczenia",
            })

        if rata1_amount:
            model.payments.append({
                "_link_case_idx": case_idx,
                "payer_type": "employer" if employer_key else "client",
                "client_key": client_key if not employer_key else None,
                "employer_key": employer_key,
                "kind": "fee",
                "amount": rata1_amount,
                "method": payment_method,
                "payment_date": rata1_date.isoformat() if rata1_date else None,
                "installment_number": 1,
                "total_installments": 2 if rata2_amount else 1,
                "notes": f"Forma: {forma_raw}",
                "import_source": "rozliczenia",
            })

        if rata2_amount:
            model.payments.append({
                "_link_case_idx": case_idx,
                "payer_type": "employer" if employer_key else "client",
                "client_key": client_key if not employer_key else None,
                "employer_key": employer_key,
                "kind": "fee",
                "amount": rata2_amount,
                "method": payment_method,
                "payment_date": rata2_date.isoformat() if rata2_date else None,
                "installment_number": 2,
                "total_installments": 2,
                "notes": f"Forma: {forma_raw}",
                "import_source": "rozliczenia",
            })

        model.stats["ROZLICZENIA"]["imported"] += 1

print(f"\n  [ROZLICZENIA.xlsx] importowanie...")
import_rozliczenia()
print(f"    imported: {model.stats['ROZLICZENIA']['imported']}, skipped: {model.stats['ROZLICZENIA']['skipped_empty']}")


# ==============================================================
# IMPORT: FAKTURY DO WYSTAWIENIA
# ==============================================================
def import_faktury():
    wb = load_workbook(find_file("ROZLICZENIA"), read_only=True, data_only=True)
    ws = wb["FAKTURY DO WYSTAWIENIA"]
    for r, row in enumerate(ws.iter_rows(values_only=True)):
        if not row or all(v is None for v in row):
            continue
        row = list(row) + [""] * max(0, 6 - len(row))
        firma = str(row[1]).strip() if row[1] else ""
        kwota = parse_amount(row[2])
        opis = str(row[3]).strip() if row[3] else ""
        wyslane_info = str(row[5]).strip() if row[5] else ""

        if not firma and not opis:
            continue

        employer_key = model.upsert_employer(firma) if firma else None

        # Parse sent info np "wysłano klient w dniu 07.01.2026"
        sent_at = None
        sent_to = None
        status = "to_issue"
        if wyslane_info:
            m = re.search(r'(\d{1,2}[.\-/]\d{1,2}[.\-/]\d{4})', wyslane_info)
            if m:
                sent_at = parse_date(m.group(1))
            if "julk" in wyslane_info.lower():
                sent_to = "Julka"
            elif "klient" in wyslane_info.lower():
                sent_to = "Klient"
            else:
                sent_to = wyslane_info[:200]
            if sent_at:
                status = "sent"

        model.invoices.append({
            "employer_key": employer_key,
            "amount": kwota or 0,
            "description": opis,
            "status": status,
            "sent_to": sent_to,
            "sent_at": sent_at.isoformat() if sent_at else None,
            "notes": wyslane_info or None,
            "import_source": "faktury_do_wystawienia",
            "import_raw": {"raw_row": [str(v) if v is not None else "" for v in row[:6]]},
        })
        model.stats["FAKTURY"]["imported"] += 1

print(f"\n  [FAKTURY DO WYSTAWIENIA] importowanie...")
import_faktury()
print(f"    imported: {model.stats['FAKTURY']['imported']}")


# ==============================================================
# IMPORT: SKŁADANE WNIOSKI (5 arkuszy)
# ==============================================================
def import_skladane_wnioski():
    path = find_file("WNIOSKI")
    doc = Document(str(path))
    # Mapowanie arkusz -> office_code + handler
    sheet_to_office = {
        "LEGNICA": "SUW_LEGNICA",
        "WAŁBRZYCH": "SUW_WALBRZYCH",
        "WROCŁAW": "DUW_WROCLAW",
        "POCZTA": None,  # wysylka pocztowa
    }
    for sheet in doc.sheets:
        sheet_name = sheet.name
        for table in sheet.tables:
            try:
                n_rows = table.num_rows
                n_cols = table.num_cols
            except Exception:
                continue

            # ODCISKI (osobiste stawiennictwo)
            if sheet_name == "ODCISKI" and table.name == "OSOBISTE STAWIENNICTWO":
                headers = [cell_val(table.cell(0, c)) for c in range(n_cols)]
                for r in range(1, n_rows):
                    row = {headers[c] if headers[c] else f"col_{c}": cell_val(table.cell(r, c)) for c in range(n_cols)}
                    first = row.get("IMIĘ", "").strip()
                    last = row.get("NAZWISKO", "").strip()
                    if not first and not last:
                        continue
                    birth = parse_date(row.get("DATA URODZENIA", ""))
                    phone = parse_phone(row.get("TELEFON", ""))
                    client_key = model.upsert_client(last, first, birth_date=birth, phone=phone)
                    opiekun = lookup_staff(row.get("OPIEKUN SPRAWY", ""))
                    termin = parse_date(row.get("Termin odcisków ", ""))
                    model.appointments.append({
                        "appointment_type": "osobiste_odciski",
                        "client_key": client_key,
                        "staff_name": opiekun,
                        "scheduled_date": termin.isoformat() if termin else None,
                        "extra": {
                            "data_wpisania": row.get("DATA WPISANIA", ""),
                            "dodatkowe_info": row.get("DODATKOWE INFORMACJE", ""),
                            "uwagi": row.get("UWAGI - co trzeba donieść na odciski", ""),
                            "transport": row.get("Jak jedzie ", ""),
                        },
                        "transport_type": row.get("Jak jedzie ", ""),
                        "import_source": "skladane_wnioski_odciski",
                    })
                    model.stats["ODCISKI"]["imported"] += 1
                continue

            # WNIOSKI ELEKTRONICZNE (bez login/hasla wg D5)
            if sheet_name == "WNIOSKI ELEKTRONICZNE" and table.name == "Tabela 1":
                headers = [cell_val(table.cell(0, c)) for c in range(n_cols)]
                for r in range(1, n_rows):
                    row = {headers[c] if headers[c] else f"col_{c}": cell_val(table.cell(r, c)) for c in range(n_cols)}
                    first = row.get("IMIĘ", "").strip()
                    last = row.get("NAZWISKO", "").strip()
                    if not first and not last:
                        continue
                    birth = parse_date(row.get("DATA URODZENIA", ""))
                    phone = parse_phone(row.get("TELEFON", ""))
                    client_key = model.upsert_client(last, first, birth_date=birth, phone=phone)
                    opiekun = lookup_staff(row.get("OPIEKUN", ""))

                    model.submissions_queue.append({
                        "client_key": client_key,
                        "office_code": None,
                        "date_notification_sent": parse_date(row.get("DATA ZŁOŻENIA WNIOSKU W KANCELARII", "")),
                        "scheduled_at": None,
                        "notes": row.get("UWAGI/INFORMACJE DODATKOWE", ""),
                        "import_source": "skladane_wnioski_elektroniczne",
                        "import_raw": {
                            **{k: v for k, v in row.items() if "PROFIL" not in k.upper()},  # D5: bez login/haslo
                            "opiekun": opiekun,
                            "termin_odciskow": row.get("TERMIN ODCISKÓW + GODZINA", ""),
                            "miejsce_odciskow": row.get("MIEJSCOWOŚĆ - ODCISKI", ""),
                            "ostatni_legalny_pobyt": row.get("DATA OSTATNIEGO DNIA LEGALNEGO POBYTU", ""),
                        },
                    })
                    model.stats["WNIOSKI ELEKTRONICZNE"]["imported"] += 1
                continue

            # LEGNICA / WAŁBRZYCH / WROCŁAW / POCZTA - standardowa struktura
            if sheet_name in sheet_to_office:
                office_code = sheet_to_office[sheet_name]
                # WAŁBRZYCH: row 0 = legenda ("LEGENDA", "WYSŁANY POCZTĄ"...),
                #           row 1 = wlasciwe headers (ZREALIZOWANE, MOS, PIO, NAZWISKO...),
                #           dane od row 2.
                # LEGNICA, WROCŁAW, POCZTA: headers w row 0, dane od row 1.
                if sheet_name == "WAŁBRZYCH":
                    header_row = 1
                    data_start = 2
                else:
                    header_row = 0
                    data_start = 1
                headers = [cell_val(table.cell(header_row, c)) for c in range(n_cols)]
                for r in range(data_start, n_rows):
                    row = {headers[c] if headers[c] else f"col_{c}": cell_val(table.cell(r, c)) for c in range(n_cols)}
                    first = row.get("IMIĘ ", "").strip() or row.get("IMIĘ", "").strip()
                    last = row.get("NAZWISKO ", "").strip() or row.get("NAZWISKO", "").strip()
                    if not first and not last:
                        continue
                    birth = parse_date(row.get("DATA URODZENIA", ""))
                    phone = parse_phone(row.get("TELEFON", ""))
                    email = row.get("EMAIL ", "").strip() or row.get("EMAIL", "").strip()
                    obyw = row.get("OBYWATELSTWO", "").strip()
                    pracodawca = row.get("PRACODAWCA", "").strip()

                    employer_key = model.upsert_employer(pracodawca) if pracodawca else None
                    client_key = model.upsert_client(
                        last, first,
                        birth_date=birth,
                        phone=phone,
                        email=email if email else None,
                        nationality=obyw if obyw else None,
                        employer_key=employer_key,
                    )

                    zrealizowane = row.get("ZREALIZOWANE", "").strip().lower()
                    is_done = zrealizowane in ("true", "tak", "yes")
                    data_wyslania = parse_date(row.get("DATA WYSŁANIA ZGŁOSZENIA", "") or row.get("DATA WYSŁANIA WNIOSKU ", ""))
                    data_stawien = row.get("DATA I GODZINA STAWIENNICTWA", "")
                    # stawien moze byc datetime
                    scheduled_at = None
                    if data_stawien:
                        try:
                            scheduled_at = datetime.strptime(data_stawien, "%Y-%m-%d %H:%M:%S")
                        except:
                            try:
                                scheduled_at = datetime.strptime(data_stawien, "%Y-%m-%d")
                            except:
                                pass

                    model.submissions_queue.append({
                        "client_key": client_key,
                        "office_code": office_code,
                        "status": "done" if is_done else "pending",
                        "mos_number": row.get("MOS", "") or row.get("MOS ", ""),
                        "pio_number": row.get("PIO", ""),
                        "ticket_number": row.get("Numer biletu", ""),
                        "date_notification_sent": data_wyslania.isoformat() if data_wyslania else None,
                        "scheduled_at": scheduled_at.isoformat() if scheduled_at else None,
                        "transport_type": row.get("PODRÓŻ", ""),
                        "notes": row.get("UWAGI", ""),
                        "import_source": f"skladane_wnioski_{sheet_name.lower()}",
                        "import_raw": row,
                    })
                    model.stats[sheet_name]["imported"] += 1

print(f"\n  [SKŁADANE WNIOSKI] importowanie...")
import_skladane_wnioski()
for k in ["LEGNICA", "WAŁBRZYCH", "WROCŁAW", "POCZTA", "ODCISKI", "WNIOSKI ELEKTRONICZNE"]:
    print(f"    [{k}] imported: {model.stats[k]['imported']}")


# ==============================================================
# PODSUMOWANIE I ZAPIS DRY-RUN
# ==============================================================
print("\n" + "=" * 70)
print("PODSUMOWANIE IMPORTU")
print("=" * 70)

summary = {
    "offices": len(model.offices),
    "office_departments": len(model.office_departments),
    "staff": len(model.staff),
    "inspectors": len(model.inspectors),
    "employers": len(model.employers),
    "clients": len(model.clients),
    "cases": len(model.cases),
    "activities": len(model.activities),
    "orphan_decisions": len(model.orphan_decisions),
    "payments": len(model.payments),
    "payment_plans": len(model.payment_plans),
    "invoices": len(model.invoices),
    "submissions_queue": len(model.submissions_queue),
    "appointments": len(model.appointments),
    "warnings": len(model.warnings),
}

for k, v in summary.items():
    print(f"  {k:25s} {v:6d}")

print(f"\nStatystyki per arkusz:")
for sheet, stats in model.stats.items():
    items = ", ".join(f"{k}={v}" for k, v in stats.items())
    print(f"  [{sheet:25s}] {items}")


# ==============================================================
# PARSE ARGS + DRY-RUN ZAPIS lub LIVE
# ==============================================================
parser = argparse.ArgumentParser()
parser.add_argument("--live", action="store_true", help="Wyslij do Supabase (inaczej tylko DRY-RUN)")
args = parser.parse_args()

if not args.live:
    # DRY-RUN - zapisz do JSON
    print("\n=== DRY-RUN: zapisuje do _dry_run/*.json ===")

    (DRY_RUN_OUT / "summary.json").write_text(
        json.dumps({
            "summary": summary,
            "stats_per_sheet": {k: dict(v) for k, v in model.stats.items()},
            "warnings_count": len(model.warnings),
        }, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

    # Zapisz kazda tabele osobno (pelne dane)
    def dump(name, data):
        path = DRY_RUN_OUT / f"{name}.json"
        path.write_text(json.dumps(data, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        print(f"  -> {name}.json ({len(data)} rekordow)")

    dump("offices", model.offices)
    dump("office_departments", model.office_departments)
    dump("staff", model.staff)
    dump("inspectors", list(model.inspectors.values()))
    dump("employers", list(model.employers.values()))
    dump("clients", list(model.clients.values()))
    dump("cases", model.cases)
    dump("activities", model.activities)
    dump("orphan_decisions", model.orphan_decisions)
    dump("payments", model.payments)
    dump("payment_plans", model.payment_plans)
    dump("invoices", model.invoices)
    dump("submissions_queue", model.submissions_queue)
    dump("appointments", model.appointments)

    if model.warnings:
        (DRY_RUN_OUT / "warnings.json").write_text(
            json.dumps(model.warnings, ensure_ascii=False, indent=2),
            encoding="utf-8"
        )
        print(f"  -> warnings.json ({len(model.warnings)} ostrzezen)")

    print("\nDRY-RUN zakonczony. Sprawdz raport w dane_od_pawla/_preview/_dry_run/summary.json")
    print("Aby wykonac prawdziwy import: py scripts/phase1-import.py --live")
else:
    print("\n=== LIVE MODE ===")
    print("Wysylanie do Supabase...")

    import requests
    import time

    URL = os.environ["SUPABASE_URL"]
    KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
    HEADERS = {
        "apikey": KEY,
        "Authorization": f"Bearer {KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }

    # Log do pliku (zeby user mogl podgladac w innym oknie)
    LOG_PATH = DRY_RUN_OUT.parent / "_live_import.log"
    log_file = open(LOG_PATH, "w", encoding="utf-8")

    def log(msg):
        line = f"[{datetime.now().strftime('%H:%M:%S')}] {msg}"
        print(line)
        log_file.write(line + "\n")
        log_file.flush()

    def batch_post(table, rows, key_field=None):
        """POST w batchach po 500. Zwraca list real UUIDs w kolejnosci rows.
        Supabase REST wymaga zeby wszystkie wiersze w batchu mialy te same klucze.
        Normalizujemy: kazdy wiersz dostaje wszystkie klucze ze wszystkich wierszy
        (brakujace -> None).
        """
        if not rows:
            return []
        # Zbierz wszystkie unikalne klucze
        all_keys = set()
        for row in rows:
            all_keys.update(row.keys())
        # Normalizuj
        normalized = []
        for row in rows:
            normalized.append({k: row.get(k) for k in all_keys})
        rows = normalized

        inserted = []
        BATCH = 500
        for i in range(0, len(rows), BATCH):
            batch = rows[i:i+BATCH]
            for attempt in range(3):
                try:
                    r = requests.post(
                        f"{URL}/rest/v1/{table}",
                        headers=HEADERS,
                        json=batch,
                        timeout=60,
                    )
                    if r.status_code >= 400:
                        log(f"  BLAD batch {i}-{i+len(batch)}: {r.status_code} {r.text[:500]}")
                        if attempt < 2:
                            time.sleep(2)
                            continue
                        raise Exception(f"Batch failed after 3 attempts: {r.status_code}")
                    inserted.extend(r.json())
                    break
                except requests.Timeout:
                    log(f"  TIMEOUT batch {i}, retry {attempt+1}")
                    if attempt == 2: raise
                    time.sleep(5)
            if i % 2000 == 0 or i + BATCH >= len(rows):
                log(f"  {table}: {i+len(batch)}/{len(rows)}")
        return inserted

    t0 = time.time()

    # ============ 1. OFFICES + DEPARTMENTS ============
    log("\n>>> Offices...")
    offices_inserted = batch_post("gmp_offices", [
        {"name": o["name"], "city": o["city"], "code": o["code"]}
        for o in model.offices
    ])
    office_code_to_id = {o["code"]: o["id"] for o in offices_inserted}
    log(f"  {len(offices_inserted)} offices")

    log(">>> Office departments...")
    dept_rows = [
        {
            "office_id": office_code_to_id[d["office_code"]],
            "code": d["code"],
            "name": d["name"],
        }
        for d in model.office_departments
    ]
    depts_inserted = batch_post("gmp_office_departments", dept_rows)
    dept_lookup = {(d["office_id"], d["code"]): d["id"] for d in depts_inserted}
    log(f"  {len(depts_inserted)} departments")

    # ============ 2. STAFF ============
    log(">>> Staff...")
    staff_rows = [
        {
            "full_name": s["full_name"],
            "role": s["role"],
            "email": s.get("email"),
            "aliases": s["aliases"],
        }
        for s in model.staff
    ]
    staff_inserted = batch_post("gmp_staff", staff_rows)
    staff_name_to_id = {s["full_name"]: s["id"] for s in staff_inserted}
    log(f"  {len(staff_inserted)} staff")

    # ============ 3. INSPECTORS ============
    log(">>> Inspectors...")
    inspector_rows = [
        {
            "full_name": i["full_name"],
            "full_name_normalized": i["full_name_normalized"],
        }
        for i in model.inspectors.values()
    ]
    insp_inserted = batch_post("gmp_inspectors", inspector_rows)
    insp_name_to_id = {i["full_name_normalized"]: i["id"] for i in insp_inserted}
    log(f"  {len(insp_inserted)} inspectors")

    # ============ 4. EMPLOYERS ============
    log(">>> Employers...")
    employer_rows = [
        {"name": e["name"], "name_normalized": e["name_normalized"]}
        for e in model.employers.values()
    ]
    emp_inserted = batch_post("gmp_employers", employer_rows)
    emp_name_to_id = {e["name_normalized"]: e["id"] for e in emp_inserted}
    log(f"  {len(emp_inserted)} employers")

    # ============ 5. CLIENTS ============
    log(">>> Clients...")
    client_rows = []
    client_keys_order = []  # zachowaj kolejnosc aby zmapowac po insercie
    for key, c in model.clients.items():
        client_keys_order.append(key)
        row = {
            "last_name": c["last_name"],
            "first_name": c["first_name"],
            "full_name_normalized": c["full_name_normalized"],
            "birth_date": c["birth_date"],
            "phone": c["phone"],
            "email": c["email"],
            "nationality": c["nationality"],
        }
        if c.get("employer_key") and c["employer_key"] in emp_name_to_id:
            row["employer_id"] = emp_name_to_id[c["employer_key"]]
        client_rows.append(row)

    cli_inserted = batch_post("gmp_clients", client_rows)
    client_key_to_id = {
        client_keys_order[i]: cli_inserted[i]["id"]
        for i in range(len(cli_inserted))
    }
    log(f"  {len(cli_inserted)} clients")

    # ============ 6. CASES ============
    log(">>> Cases...")
    case_rows = []
    case_indices_order = []
    for idx, c in enumerate(model.cases):
        case_indices_order.append(idx)
        row = {
            "case_number": c.get("case_number"),
            "znak_sprawy": c.get("znak_sprawy"),
            "status": c.get("status", "aktywna"),
            "status_notes": c.get("status_notes"),
            "stage": c.get("stage"),
            "stage_notes": c.get("stage_notes"),
            "kind": c.get("kind", "nowa_sprawa"),
            "case_type": c.get("case_type"),
            "case_type_notes": c.get("case_type_notes"),
            "submission_method": c.get("submission_method"),
            "category": c.get("category"),
            "date_accepted": c.get("date_accepted"),
            "date_submitted": c.get("date_submitted"),
            "date_joined": c.get("date_joined"),
            "date_transfer_request": c.get("date_transfer_request"),
            "date_transferred": c.get("date_transferred"),
            "fee_amount": c.get("fee_amount"),
            "fee_notes": c.get("fee_notes"),
            "paragon": c.get("paragon", False),
            "extra_notes": c.get("extra_notes"),
            "import_source": c.get("import_source"),
            "import_source_row": c.get("import_source_row"),
            "import_raw": c.get("import_raw"),
        }
        # Relacje
        if c.get("client_key") and c["client_key"] in client_key_to_id:
            row["client_id"] = client_key_to_id[c["client_key"]]
        if c.get("employer_key") and c["employer_key"] in emp_name_to_id:
            row["employer_id"] = emp_name_to_id[c["employer_key"]]
        if c.get("staff_name") and c["staff_name"] in staff_name_to_id:
            row["assigned_to"] = staff_name_to_id[c["staff_name"]]
        if c.get("inspector_key") and c["inspector_key"] in insp_name_to_id:
            row["inspector_id"] = insp_name_to_id[c["inspector_key"]]
        if c.get("office_code") and c["office_code"] in office_code_to_id:
            row["office_id"] = office_code_to_id[c["office_code"]]
            if c.get("department_code"):
                dept_key = (office_code_to_id[c["office_code"]], c["department_code"])
                if dept_key in dept_lookup:
                    row["department_id"] = dept_lookup[dept_key]
        case_rows.append(row)

    cases_inserted = batch_post("gmp_cases", case_rows)
    case_idx_to_id = {
        case_indices_order[i]: cases_inserted[i]["id"]
        for i in range(len(cases_inserted))
    }
    # Mapowanie client_key -> lista case_id (do activities decyzji)
    client_to_case_ids = defaultdict(list)
    for idx, case in enumerate(model.cases):
        if idx in case_idx_to_id and case.get("client_key"):
            client_to_case_ids[case["client_key"]].append(case_idx_to_id[idx])
    log(f"  {len(cases_inserted)} cases")

    # ============ 7. CASE ACTIVITIES (odebrane decyzje matched) ============
    log(">>> Activities (odebrane decyzje)...")
    act_rows = []
    for a in model.activities:
        match_key = a.pop("_match_client_key", None)
        if match_key and match_key in client_to_case_ids:
            case_ids = client_to_case_ids[match_key]
            # Dolaczamy do pierwszej sprawy klienta
            act_rows.append({
                "case_id": case_ids[0],
                "activity_type": a["activity_type"],
                "content": a["content"],
                "metadata": a.get("metadata"),
            })
    acts_inserted = batch_post("gmp_case_activities", act_rows) if act_rows else []
    log(f"  {len(acts_inserted)} activities")

    # ============ 8. ORPHAN DECISIONS ============
    log(">>> Orphan decisions...")
    orphan_rows = [
        {
            "raw_full_name": o["raw_full_name"],
            "raw_birth_date": o["raw_birth_date"],
            "date_issued": o["date_issued"],
            "date_received": o["date_received"],
            "date_delivered_to_client": o["date_delivered_to_client"],
            "extra_notes": o.get("extra_notes"),
        }
        for o in model.orphan_decisions
    ]
    orph_inserted = batch_post("gmp_orphan_decisions", orphan_rows) if orphan_rows else []
    log(f"  {len(orph_inserted)} orphan decisions")

    # ============ 9. PAYMENT PLANS ============
    log(">>> Payment plans...")
    plan_rows = []
    for p in model.payment_plans:
        case_idx = p.get("_link_case_idx")
        if case_idx not in case_idx_to_id:
            continue
        row = {
            "case_id": case_idx_to_id[case_idx],
            "total_amount": p["total_amount"],
            "payer_type": p["payer_type"],
            "installments_planned": p.get("installments_planned", 1),
        }
        if p.get("client_key") and p["client_key"] in client_key_to_id:
            row["client_id"] = client_key_to_id[p["client_key"]]
        if p.get("employer_key") and p["employer_key"] in emp_name_to_id:
            row["employer_id"] = emp_name_to_id[p["employer_key"]]
        plan_rows.append(row)
    plans_inserted = batch_post("gmp_payment_plans", plan_rows) if plan_rows else []
    log(f"  {len(plans_inserted)} payment plans")

    # ============ 10. PAYMENTS ============
    log(">>> Payments...")
    pay_rows = []
    for p in model.payments:
        case_idx = p.get("_link_case_idx")
        if case_idx not in case_idx_to_id:
            continue
        row = {
            "case_id": case_idx_to_id[case_idx],
            "payer_type": p["payer_type"],
            "kind": p.get("kind", "fee"),
            "amount": p["amount"],
            "method": p.get("method"),
            "payment_date": p.get("payment_date"),
            "installment_number": p.get("installment_number"),
            "total_installments": p.get("total_installments"),
            "notes": p.get("notes"),
            "import_source": p.get("import_source"),
        }
        if p.get("client_key") and p["client_key"] in client_key_to_id:
            row["client_id"] = client_key_to_id[p["client_key"]]
        if p.get("employer_key") and p["employer_key"] in emp_name_to_id:
            row["employer_id"] = emp_name_to_id[p["employer_key"]]
        pay_rows.append(row)
    pays_inserted = batch_post("gmp_payments", pay_rows) if pay_rows else []
    log(f"  {len(pays_inserted)} payments")

    # ============ 11. INVOICES ============
    log(">>> Invoices...")
    inv_rows = []
    for inv in model.invoices:
        row = {
            "amount": inv["amount"] or 0,
            "description": inv.get("description"),
            "status": inv.get("status", "to_issue"),
            "sent_to": inv.get("sent_to"),
            "sent_at": inv.get("sent_at"),
            "notes": inv.get("notes"),
            "import_source": inv.get("import_source"),
            "import_raw": inv.get("import_raw"),
        }
        if inv.get("employer_key") and inv["employer_key"] in emp_name_to_id:
            row["employer_id"] = emp_name_to_id[inv["employer_key"]]
        if inv.get("client_key") and inv["client_key"] in client_key_to_id:
            row["client_id"] = client_key_to_id[inv["client_key"]]
        case_idx = inv.get("_link_case_idx")
        if case_idx is not None and case_idx in case_idx_to_id:
            row["case_id"] = case_idx_to_id[case_idx]
        inv_rows.append(row)
    inv_inserted = batch_post("gmp_invoices", inv_rows) if inv_rows else []
    log(f"  {len(inv_inserted)} invoices")

    # ============ 12. SUBMISSIONS_QUEUE ============
    log(">>> Submissions queue...")
    sub_rows = []
    for s in model.submissions_queue:
        row = {
            "status": s.get("status", "pending"),
            "mos_number": s.get("mos_number"),
            "pio_number": s.get("pio_number"),
            "ticket_number": s.get("ticket_number"),
            "date_notification_sent": s.get("date_notification_sent"),
            "scheduled_at": s.get("scheduled_at"),
            "transport_type": s.get("transport_type"),
            "notes": s.get("notes"),
            "import_source": s.get("import_source"),
            "import_raw": s.get("import_raw"),
        }
        if s.get("client_key") and s["client_key"] in client_key_to_id:
            row["client_id"] = client_key_to_id[s["client_key"]]
        if s.get("office_code") and s["office_code"] in office_code_to_id:
            row["office_id"] = office_code_to_id[s["office_code"]]
        sub_rows.append(row)
    subs_inserted = batch_post("gmp_submissions_queue", sub_rows) if sub_rows else []
    log(f"  {len(subs_inserted)} submissions_queue")

    # ============ 13. APPOINTMENTS ============
    log(">>> Appointments...")
    app_rows = []
    for a in model.appointments:
        row = {
            "appointment_type": a.get("appointment_type", "konsultacja"),
            "scheduled_date": a.get("scheduled_date"),
            "transport_type": a.get("transport_type"),
        }
        # Stara tabela gmp_appointments ma wiele NOT NULL fields (lawyer_id, client_name)
        # Uzywamy tylko nowych kolumn - stare ustawiamy na defaulty
        # Potrzebujemy lawyer_id - NOT NULL w oryginalnej migracji!
        if a.get("staff_name") and a["staff_name"] in staff_name_to_id:
            row["staff_id"] = staff_name_to_id[a["staff_name"]]
        if a.get("client_key") and a["client_key"] in client_key_to_id:
            row["client_id"] = client_key_to_id[a["client_key"]]
        case_idx = a.get("_link_case_idx")
        if case_idx is not None and case_idx in case_idx_to_id:
            row["case_id"] = case_idx_to_id[case_idx]
        app_rows.append(row)

    # gmp_appointments z oryginalnej migracji ma NOT NULL lawyer_id + scheduled_date + scheduled_time
    # Tworzymy NOWA tabele gmp_appointments_crm zamiast borykac sie ze starą
    # Na razie pomijam import appointments - zrobimy w Fazie 4 gdy budujemy UI
    log(f"  SKIP: appointments ({len(app_rows)} rekordow) - stara tabela gmp_appointments ma NOT NULL constraints. Zaimplementujemy w Fazie 4.")

    t_end = time.time()
    log(f"\n=== LIVE IMPORT ZAKONCZONY w {t_end-t0:.1f}s ===")

    # Weryfikacja finalna - zliczamy z bazy
    log("\nWeryfikacja liczb w bazie:")
    for table in [
        "gmp_offices", "gmp_office_departments", "gmp_staff", "gmp_inspectors",
        "gmp_employers", "gmp_clients", "gmp_cases", "gmp_case_activities",
        "gmp_orphan_decisions", "gmp_payment_plans", "gmp_payments",
        "gmp_invoices", "gmp_submissions_queue"
    ]:
        r = requests.get(
            f"{URL}/rest/v1/{table}?select=count",
            headers={**HEADERS, "Prefer": "count=exact", "Range": "0-0"},
            timeout=30,
        )
        cr = r.headers.get("content-range", "")
        count = cr.split("/")[-1] if cr else "?"
        log(f"  {table:30s} {count}")

    log_file.close()
